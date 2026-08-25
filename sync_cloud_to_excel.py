# -*- coding: utf-8 -*-
"""
PS5 - Sync platform edits (GitHub Pages) -> PS-5 COMPLETIONS DPR SUMMERY Excel.

Pulls https://mohamedgawad1.github.io/PS5-COMPLETION-PLATFORM/platform_state.json
and applies every platform edit to the DPR SUMMERY workbook with EXACT mapping:

  PUNCH LIST -> sheet 'DETAILED PUNCH LIST'  (id col 'Punchlist ID')
  ITR LIST   -> sheet 'DETAILED ITR LIST'    (id col 'Task ID')
  RFC PROGRESS -> sheet 'RFC PROGRESS'       (id = text before ' - ' in col A,
                                              data starts row 4)

Column map is fixed below (platform name -> exact Excel header). Nothing is
written into formula cells (BALANCE / % / ITRs / CLOSED totals are formulas and
recalculate automatically). Notes go to a 'PLATFORM NOTES' column that is
created at the far right if missing. Row colors become the fill of the ID cell.

Usage:
  python sync_cloud_to_excel.py            # one shot
  python sync_cloud_to_excel.py --watch    # repeat every 5 minutes
  python sync_cloud_to_excel.py --force    # re-apply even if state unchanged
"""
import hashlib
import json
import os
import re
import shutil
import sys
import time
import urllib.request
import zipfile

import openpyxl
from openpyxl.styles import Font, PatternFill

HOME = os.path.join(os.path.expanduser('~'), 'Downloads')
DL_SUB = os.path.join(HOME, 'PS5 - CPP AGI Completion Progress Dashboard_files')
HERE = os.path.dirname(os.path.abspath(__file__))

STATE_URL = ('https://raw.githubusercontent.com/Mohamedgawad1/'
             'PS5-COMPLETION-PLATFORM/main/platform_state.json')
SEEN_FILE = os.path.join(HERE, '_platform_sync_seen.json')
REPORT_FILE = os.path.join(HERE, '_platform_sync_report.txt')

# ---------------------------------------------------------------- mapping --
# platform column  -> exact Excel header text
PUNCH_MAP = {
    'TAG': 'Asset (Name/Tag)',
    'CAT': 'CAT',
    'DISC': 'Discipline (Name)',
    'DESCRIPTION': 'Description',
    'STATUS': 'Status',
    'CLOSING DATE': 'Workflow - Closing Date',
}
ITR_MAP = {
    'TAG': 'Asset - Tag',
    'DISC': 'Discipline',
    'TASK TYPE': 'Task Type (Name)',
    'ASSET DESCRIPTION': 'Asset - Description',
    'STATE': 'Task State',
    'CLOSING DATE': 'Closing Date',
}
# RFC PROGRESS sheet: positional (1-based) columns, verified against the file
RFC_ID_COL = 1          # 'PS5-01-01 - description'
RFC_DATA_ROW = 4        # first data row
RFC_MAP = {
    'Priority': 2,
    'RFC BHMPS': 3,
    'RFC EIT': 4,
    'Baseline': 5,
    'Recovery': 6,
    'SIGNED': 7,
    'Milestone': 8,
    'CPP-1': 50,
    'EIT': 51,
    'EACOP': 52,
    'REMARK EACOP': 53,
    'REMARK CPP-EIT': 54,
    'REMARK CPP-1': 55,
    'STATUS': None,      # resolved dynamically -> 'WALKDOWN STATUS' column
}
# discipline letter -> (TOTAL col, CLOSED col)
RFC_DISC_COLS = {'B': (10, 11), 'E': (14, 15), 'H': (18, 19), 'I': (22, 23),
                 'M': (26, 27), 'P': (30, 31), 'S': (34, 35), 'T': (38, 39)}
# platform columns that are derived in Excel (formulas) - never written
RFC_DERIVED = {'ITRs', 'CLOSED', 'BALANCE'}

NOTE_HEADER = 'PLATFORM NOTES'
TARGET_SHEET = {
    'PUNCH LIST': ('DETAILED PUNCH LIST', 'Punchlist ID', PUNCH_MAP),
    'ITR LIST': ('DETAILED ITR LIST', 'Task ID', ITR_MAP),
}


def n(v):
    return '' if v is None else str(v).strip()


def fetch_state(path=None):
    if path:
        try:
            with open(path, encoding='utf-8') as f:
                return json.load(f)
        except Exception as e:
            print(f'[cloud] state file read failed: {e}')
            return None
    try:
        req = urllib.request.Request(STATE_URL, headers={'User-Agent': 'ps5-sync'})
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode('utf-8'))
    except Exception as e:
        print(f'[cloud] download failed: {e}')
        return None


def load_seen():
    try:
        with open(SEEN_FILE, encoding='utf-8') as f:
            return json.load(f).get('hash')
    except Exception:
        return None


def save_seen(h):
    with open(SEEN_FILE, 'w', encoding='utf-8') as f:
        json.dump({'hash': h, 'time': time.strftime('%Y-%m-%d %H:%M:%S')}, f)


DATE_RE = re.compile(r'PS-5\s*COMPLETIONS\s*DPR\s*SUMMERY\s*-\s*'
                     r'(\d{1,2})-(\d{1,2})-(\d{2,4})\.xlsx$', re.I)


def _date_key(m):
    """(year, month, day) tuple sortable as strings."""
    dd, mm, yy = m.group(1), m.group(2), m.group(3)
    yy = yy if len(yy) == 4 else '20' + yy
    return (yy, mm.zfill(2), dd.zfill(2))


def find_target():
    """Newest DPR SUMMERY by the DATE SUFFIX in its name (DD-MM-YY),
    e.g. 'PS-5 COMPLETIONS DPR SUMMERY -25-08-26.xlsx'.
    Only the date changes daily - that suffix decides, not file mtime."""
    best = None
    seen = set()
    for folder in (HOME, DL_SUB):
        if not os.path.isdir(folder):
            continue
        for f in os.listdir(folder):
            b = f.upper()
            if (f.startswith('~$') or not f.lower().endswith('.xlsx')
                    or 'BACKUP' in b):
                continue
            m = DATE_RE.search(f)
            if not m:
                continue
            p = os.path.join(folder, f)
            key = os.path.realpath(p)
            if key in seen:
                continue
            seen.add(key)
            cand = (_date_key(m), os.path.getmtime(p), p)
            if best is None or cand[:2] > best[:2]:
                best = cand
    return best[2] if best else None


def header_index(ws, title):
    """Exact (case/space-insensitive) header lookup in row 1."""
    want = re.sub(r'\s+', ' ', n(title)).upper()
    for c in range(1, ws.max_column + 1):
        if re.sub(r'\s+', ' ', n(ws.cell(1, c).value)).upper() == want:
            return c
    return None


def ensure_note_col(ws):
    c = header_index(ws, NOTE_HEADER)
    if c:
        return c
    c = ws.max_column + 1
    ws.cell(1, c).value = NOTE_HEADER
    return c


def is_formula(cell):
    v = cell.value
    if isinstance(v, str) and v.startswith('='):
        return True
    return type(v).__name__ == 'ArrayFormula'


def build_rfc_ids(ws):
    ids = {}
    for r in range(RFC_DATA_ROW, ws.max_row + 1):
        raw = n(ws.cell(r, RFC_ID_COL).value)
        if not raw:
            continue
        sid = raw.split(' - ')[0].strip().upper()
        if sid and sid not in ids:
            ids[sid] = r
    return ids


def apply_fill(ws, row, col, color):
    cell = ws.cell(row, col)
    if color:
        argb = color if len(color) == 8 else 'FF' + color.lstrip('#')
        cell.fill = PatternFill(start_color=argb, end_color=argb,
                                fill_type='solid')


def index_ids(ws, idc, start=2):
    """Map ID -> row. Stops after a long run of empty rows (formatted-but-empty
    sheets like DETAILED ITR LIST report max_row = 1048576)."""
    rows_by_id = {}
    empty = 0
    for r in range(start, ws.max_row + 1):
        rid = n(ws.cell(r, idc).value)
        if rid:
            rows_by_id[rid.upper()] = r
            empty = 0
        else:
            empty += 1
            if empty > 300:
                break
    return rows_by_id


def _num(v):
    """Numeric value of a possibly-formula cell from a data_only workbook."""
    if v is None or isinstance(v, str):
        try:
            return float(str(v).strip() or 0)
        except ValueError:
            return 0.0
    return float(v)


def sync_prc_status(src, wb, rep):
    """Make 'PRC STATUS' reflect platform page 7 (COMPLETE):
    effective TOTAL % + RFC SIGNED per subsystem, green mark at 100%."""
    if 'PRC STATUS' not in wb.sheetnames:
        return
    wbd = openpyxl.load_workbook(src, data_only=True)
    try:
        rws = wbd['RFC PROGRESS']
        eff = {}
        for kfull, r in index_ids(rws, RFC_ID_COL,
                                  start=RFC_DATA_ROW).items():
            sid = kfull.split(' - ')[0].strip()
            sT = sC = 0.0
            for tc, cc in RFC_DISC_COLS.values():
                sT += _num(rws.cell(r, tc).value)
                sC += _num(rws.cell(r, cc).value)
            eff[sid] = (round(sC / sT * 100) if sT > 0 else None,
                        n(rws.cell(r, 7).value))
    finally:
        wbd.close()

    pws = wb['PRC STATUS']

    def find_or_make(title, row=1):
        want = re.sub(r'\s+', ' ', title).upper()
        for c in range(1, pws.max_column + 1):
            if re.sub(r'\s+', ' ', n(pws.cell(row, c).value)).upper() == want:
                return c
        c = pws.max_column + 1
        pws.cell(row, c).value = title
        return c

    c_tot = find_or_make('TOTAL %')
    c_sig = find_or_make('RFC SIGNED')

    empty = 0
    for r in range(3, pws.max_row + 1):
        v = n(pws.cell(r, 1).value)
        if not v:
            empty += 1
            if empty > 300:
                break
            continue
        empty = 0
        sid = v.split(' - ')[0].strip()
        tot, signed = eff.get(sid, (None, ''))
        cell = pws.cell(r, c_tot)
        if tot is not None and n(cell.value) != str(tot):
            cell.value = tot
            cell.font = Font(size=9)
            rep['prc'] += 1
        if tot is not None and tot >= 100:
            cur = getattr(getattr(cell.fill, 'start_color', None), 'rgb', '')
            if cur != 'FFC6EFCE':
                apply_fill(pws, r, c_tot, 'FFC6EFCE')
                rep['prc'] += 1
        scell = pws.cell(r, c_sig)
        sv = signed.strip() if signed else ''
        if sv and n(scell.value) != sv:
            scell.value = sv
            scell.font = Font(size=9)
            rep['prc'] += 1


def main():
    args = sys.argv[1:]
    force = '--force' in args
    watch = '--watch' in args
    state_file = None
    target = None
    if '--state' in args:
        state_file = args[args.index('--state') + 1]
    if '--target' in args:
        target = args[args.index('--target') + 1]
    while True:
        try:
            run_once(force, state_file, target)
        except Exception as e:
            print('[sync] error:', e)
        if not watch:
            return 0
        time.sleep(300)


def run_once(force, state_file=None, target=None):
    global SEEN_FILE
    st = fetch_state(state_file)
    if st is None:
        return
    if state_file:
        force = True
    blob = json.dumps(st, sort_keys=True, ensure_ascii=False)
    h = hashlib.sha256(blob.encode()).hexdigest()
    if h == load_seen() and not force:
        print('[sync] cloud state unchanged - nothing to do')
        return

    src = target or find_target()
    if not src:
        print('[sync] no DPR SUMMERY workbook found!')
        return
    print(f'[sync] target : {os.path.basename(src)}')

    cells = st.get('cells', {}) or {}
    notes = st.get('notes', {}) or {}
    colors = st.get('colors', {}) or {}

    backup = os.path.join(os.path.dirname(src),
                          os.path.splitext(os.path.basename(src))[0]
                          + ' -BACKUP.xlsx')
    shutil.copy2(src, backup)

    try:
        wb = openpyxl.load_workbook(src)
    except Exception as e:
        print(f'[sync] cannot open "{os.path.basename(src)}" ({e}) - '
              f'close Excel / check the file. NOTHING was changed.')
        return
    rep = {'written': 0, 'notes': 0, 'colors': 0, 'skipped_formula': [],
           'skipped_derived': [], 'no_row': [], 'no_col': []}

    # ---- detailed sheets -------------------------------------------------
    for psheet, (tgt, idh, cmap) in TARGET_SHEET.items():
        has_edits = (cells.get(psheet) or notes.get(psheet)
                     or colors.get(psheet))
        if tgt not in wb.sheetnames or not has_edits:
            continue
        ws = wb[tgt]
        idc = header_index(ws, idh)
        if not idc:
            rep['no_col'].append(f'{tgt}: ID column {idh!r} missing')
            continue
        rows_by_id = index_ids(ws, idc)
        colmap = {}
        for pcol, etitle in cmap.items():
            ec = header_index(ws, etitle)
            if ec:
                colmap[pcol] = ec
            else:
                rep['no_col'].append(f'{tgt}: {etitle!r} missing')
        note_c = ensure_note_col(ws) if notes.get(psheet) else None

        for rid, ed in (cells.get(psheet, {}) or {}).items():
            r = rows_by_id.get(n(rid).upper())
            if not r:
                rep['no_row'].append(f'{psheet}/{rid}')
                continue
            for pcol, val in ed.items():
                ec = colmap.get(pcol)
                if not ec:
                    rep['no_col'].append(f'{psheet}: {pcol} unmapped')
                    continue
                cell = ws.cell(r, ec)
                if is_formula(cell):
                    rep['skipped_formula'].append(f'{tgt}!{cell.coordinate}')
                    continue
                cell.value = val
                cell.font = Font(size=9)
                rep['written'] += 1
        for rid, val in (notes.get(psheet, {}) or {}).items():
            r = rows_by_id.get(n(rid).upper())
            if not r:
                rep['no_row'].append(f'{psheet}-note/{rid}')
                continue
            ws.cell(r, note_c).value = val
            ws.cell(r, note_c).font = Font(size=9)
            rep['notes'] += 1
        for rid, col in (colors.get(psheet, {}) or {}).items():
            r = rows_by_id.get(n(rid).upper())
            if r:
                apply_fill(ws, r, 1, col)
                rep['colors'] += 1

    # ---- RFC PROGRESS ----------------------------------------------------
    sh = 'RFC PROGRESS'
    if sh in wb.sheetnames and (cells.get(sh) or notes.get(sh)
                                or colors.get(sh)):
        ws = wb[sh]
        # platform sid = text before ' - ' in column A
        ids = {k.split(' - ')[0].strip(): v
               for k, v in index_ids(ws, RFC_ID_COL,
                                     start=RFC_DATA_ROW).items()}
        walk_holder = [None]

        def get_walk():
            if walk_holder[0] is None:
                c = None
                for cc in range(56, ws.max_column + 1):
                    if re.sub(r'\s+', ' ', n(ws.cell(3, cc).value)).upper() \
                            == 'WALKDOWN STATUS':
                        c = cc
                        break
                if c is None:
                    c = max(ws.max_column, 55) + 1
                    ws.cell(3, c).value = 'WALKDOWN STATUS'
                walk_holder[0] = c
            return walk_holder[0]

        def put(r, cidx, val):
            cell = ws.cell(r, cidx)
            if is_formula(cell):
                rep['skipped_formula'].append(f'{sh}!{cell.coordinate}')
                return
            cell.value = val
            cell.font = Font(size=9)
            rep['written'] += 1

        for rid, ed in (cells.get(sh, {}) or {}).items():
            r = ids.get(n(rid).upper())
            if not r:
                rep['no_row'].append(f'{sh}/{rid}')
                continue
            for pcol, val in ed.items():
                mkey = re.match(r'^([A-T]) (TOTAL|CLOSED)$', pcol)
                if mkey and mkey.group(1) in RFC_DISC_COLS:
                    tc, cc = RFC_DISC_COLS[mkey.group(1)]
                    put(r, tc if mkey.group(2) == 'TOTAL' else cc, val)
                elif pcol in RFC_DERIVED:
                    rep['skipped_derived'].append(f'{sh}:{pcol} ({rid})')
                elif pcol in RFC_MAP:
                    cidx = RFC_MAP[pcol]
                    if cidx is None:
                        cidx = get_walk()
                    put(r, cidx, val)
                else:
                    rep['no_col'].append(f'{sh}: {pcol} unmapped')
        for rid, val in (notes.get(sh, {}) or {}).items():
            r = ids.get(n(rid).upper())
            if not r:
                rep['no_row'].append(f'{sh}-note/{rid}')
                continue
            nc = ensure_note_col(ws)
            ws.cell(r, nc).value = val
            ws.cell(r, nc).font = Font(size=9)
            rep['notes'] += 1
        for rid, col in (colors.get(sh, {}) or {}).items():
            r = ids.get(n(rid).upper())
            if r:
                apply_fill(ws, r, 1, col)
                rep['colors'] += 1

    # ---- PRC STATUS = platform page 7 ------------------------------------
    rep.setdefault('prc', 0)
    sync_prc_status(src, wb, rep)

    # ---- nothing changed? do not touch the workbook at all ---------------
    if not (rep['written'] or rep['notes'] or rep['colors'] or rep['prc']):
        wb.close()
        save_seen(h) if not state_file else None
        print('[sync] no effective changes - workbook untouched')
        return

    # ---- atomic save: tmp file -> validate -> replace --------------------
    tmpf = src + '.saving.tmp'
    ok = False
    for attempt in range(15):
        try:
            wb.save(tmpf)
            ok = True
            break
        except PermissionError:
            print(f'[wait] close "{os.path.basename(src)}" in Excel '
                  f'({attempt + 1}/15)...')
            time.sleep(2)
    wb.close()
    if not ok:
        if os.path.exists(tmpf):
            os.remove(tmpf)
        print('[sync] file locked - nothing changed, try again later')
        return
    try:
        _z = zipfile.ZipFile(tmpf)
        _bad = _z.testzip()
        _n = len(_z.namelist())
        _z.close()
        if _bad is not None or _n < 10:
            raise RuntimeError(f'invalid archive ({_n} entries)')
    except Exception as e:
        if os.path.exists(tmpf):
            os.remove(tmpf)
        print(f'[sync] post-save validation FAILED ({e}) - '
              f'original file untouched')
        return
    os.replace(tmpf, src)

    save_seen(h) if not state_file else None
    lines = [
        f"sync {time.strftime('%Y-%m-%d %H:%M:%S')} -> {os.path.basename(src)}",
        f"written={rep['written']} notes={rep['notes']} "
        f"colors={rep['colors']} prc={rep['prc']}"]
    if rep['skipped_derived']:
        lines.append('derived(auto in Excel, not written): '
                     + ', '.join(rep['skipped_derived'][:20]))
    if rep['skipped_formula']:
        lines.append('formula cells protected: '
                     + ', '.join(rep['skipped_formula'][:20]))
    if rep['no_row']:
        lines.append('rows NOT found: ' + ', '.join(rep['no_row'][:30]))
    if rep['no_col']:
        lines.append('columns missing/unmapped: '
                     + ', '.join(sorted(set(rep['no_col']))[:20]))
    text = '\n'.join(lines)
    print(text)
    with open(REPORT_FILE, 'w', encoding='utf-8') as f:
        f.write(text + '\n')


if __name__ == '__main__':
    raise SystemExit(main())
