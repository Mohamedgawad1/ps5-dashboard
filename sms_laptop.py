import openpyxl, schedule, time, datetime, subprocess, os

RECIPIENTS = [
    "+255684064292",
    "+255677114429",
    "+255674643758",
    "+255699065843",
]

TRACKER = r"C:\Users\mylap\OneDrive\Desktop\dashboard\PS5 Master tracker EIT Combined.xlsx"
OV_FILE = r"C:\Users\mylap\OneDrive\Desktop\dashboard\ovTasks_TestsPlanned_1369.xlsx"
ADB = r"C:\adb\platform-tools\adb"
PHONE = "swjrci9xythaizuk"

def send_sms(number, msg):
    safe_body = msg.replace('"', "'").replace("&", "and")
    subprocess.run(
        f'{ADB} -s {PHONE} shell am start -a android.intent.action.SENDTO -d sms:{number} -e sms_body "{safe_body}"',
        shell=True
    )
    print(f"  Opened for {number}")

def get_today_counts():
    today = datetime.date.today()
    wb = openpyxl.load_workbook(OV_FILE, data_only=True)
    ws = wb.active
    tc = ts = 0
    for r in range(2, ws.max_row + 1):
        cd = ws.cell(r, 22).value
        st = str(ws.cell(r, 28).value or "").lower()
        if cd:
            if isinstance(cd, datetime.datetime): d = cd.date()
            elif isinstance(cd, str):
                try: d = datetime.datetime.strptime(cd.split()[0], '%Y-%m-%d').date()
                except: d = None
            else: d = None
            if d == today:
                if "close" in st: tc += 1
                elif "submitt" in st: ts += 1
    wb.close()
    return tc, ts

def get_totals():
    wb = openpyxl.load_workbook(TRACKER, data_only=True)
    ws = wb["ITR Tracker"]
    c = ws.cell(7, 2).value or 0
    n = ws.cell(9, 2).value or 0
    wb.close()
    return c, n

def send_status():
    now = datetime.datetime.now()
    tc, ts = get_today_counts()
    c, n = get_totals()
    msg = f"ITR: today close {tc} submitt {ts}\nITR TOTAL CLOSE {c} OPEN {n}"
    print(f"\n[{now.strftime('%H:%M')}] Opening Messages for {len(RECIPIENTS)} recipients...")
    print(msg)
    print("  -> Tap SEND on phone for each one!")
    print("  -> Going to next in 8 seconds after each...")
    for r in RECIPIENTS:
        send_sms(r, msg)
        time.sleep(8)

for t in ["13:00", "16:00"]:
    schedule.every().day.at(t).do(send_status)

print("SMS Dashboard running on laptop. Opens Messages at 13:00, 16:00")
print("Waiting for next scheduled time...")
print("  (Or run send_status() manually to test now)")
